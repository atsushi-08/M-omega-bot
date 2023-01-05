import discord
from discord.ext import commands
import random
from openpyxl import load_workbook
import json

class Kit(commands.Cog):
    def __init__(self,bot :commands.Bot) -> None:
        self.bot = bot

    @commands.command()
    async def 抽歌(self,ctx,game,lower,*upper):
        if len(upper) == 0:
            upper = [float(lower)]
        if game == '中二':
            wb = load_workbook('中二定數.xlsx')
            list = []
            for sheet in wb:
                for row in sheet.iter_rows(min_col=1,max_col=4):
                    if row[3].value >= float(lower) and row[3].value <= float(upper[0]):
                        list.append([row[0].value,row[1].value])
            if len(list) == 0:
                await ctx.send('無符合條件的歌')
            else:
                song = random.choice(list)
                await ctx.send(f"{song[0]} {song[1]}")
                    
        if game == 'mai':
            wb = load_workbook('mai定數.xlsx')
            list = []
            for sheet in wb:
                for row in sheet.iter_rows(min_col=1,max_col=5):
                    if row[4].value >= float(lower) and row[4].value <= float(upper[0]):
                        list.append([f'{row[0].value} ({row[2].value}譜)',row[3].value])
            if len(list) == 0:
                await ctx.send('無符合條件的歌')
            else:
                song = random.choice(list)
                await ctx.send(f"{song[0]} {song[1]}")
                    
    @commands.command()
    async def 簡稱(self,ctx,method,abbr,song):
        with open('abbreviation.json','r',encoding='utf-8') as j:
            data = json.load(j)
            j.close

        if method == '增加':
            if abbr in data:
                await ctx.send(f'這個簡稱已經被{data[abbr]}用過了')
                return
            else:
                data[abbr] = song
                with open('abbreviation.json','w',encoding='utf-8') as j:
                    json.dump(data,j,ensure_ascii=False,indent=4)
                    j.close
                await ctx.send(f'已為{song}設定簡稱{abbr}')
        elif method == '移除':
            if abbr in data:
                del data[abbr]
                with open('abbreviation.json','w',encoding='utf-8') as j:
                    json.dump(data,j,ensure_ascii=False,indent=4)
                    j.close
                await ctx.send('已移除簡稱')
            else:
                await ctx.send(f'找不到使用這個簡稱的歌')
                return
        else:
            await ctx.send('目前只有增加跟移除兩個功能喔')

    @commands.command()
    async def 隨機段位(self,ctx,game,level):
        wb = load_workbook('中二定數.xlsx')
        dif = {
            '四段' : [[13.5,13.9],[14.0,14.4],[14.5,14.9]],
            '五段' : [[14.5,14.9],[14.5,14.9],[14.5,14.9]],
            '無限段' : [[14.0,14.4],[14.5,14.9],[15.0,15.4]]
        }
        if level in dif:
            first = []
            second = []
            third = []
            for sheet in wb:
                for row in sheet.iter_rows(min_col=1,max_col=4):
                    if row[3].value >= dif[level][0][0] and row[3].value <= dif[level][0][1]:
                        first.append([row[0].value,row[1].value])
                    if row[3].value >= dif[level][1][0] and row[3].value <= dif[level][1][1]:
                        second.append([row[0].value,row[1].value])
                    if row[3].value >= dif[level][2][0] and row[3].value <= dif[level][2][1]:
                        third.append([row[0].value,row[1].value])
            songs = [random.choice(first),random.choice(second),random.choice(third)]
            await ctx.send(f'模擬隨機{level}的結果:\n第一首{songs[0][0]} {songs[0][1]}\n第二首{songs[1][0]} {songs[1][1]}\n第三首{songs[2][0]} {songs[2][1]}\n')
        else:
            await ctx.send('無法辨別的段位')


async def setup(bot:commands.Bot) -> None:
    await bot.add_cog(Kit(bot))