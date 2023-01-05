from openpyxl import load_workbook
import discord
from discord.ext import commands
import json

class Chumai(commands.Cog):
    def __init__(self,bot :commands.Bot) -> None:
        self.bot = bot
    
    @commands.command()
    async def 中二(self,ctx,name,dif,*score):
        def rating(base,sc):
            if sc>1009000:
                return round(base+2.15,2)
            elif sc>1007500:
                return round(base+2+(sc-1007500)/10000-0.005,2)
            elif sc>1005000:
                return round(base+1.5+(sc-1005000)/5000-0.005,2)
            elif sc>1000000:
                return round(base+1+(sc-1000000)/10000-0.005,2)
            else:
                return round(base+(sc-975000)/25000-0.005,2)

        wb = load_workbook('中二定數.xlsx')
        find = 0
        with open('abbreviation.json','r',encoding='utf-8') as j:
            data = json.load(j)
            j.close
        if name in data:
            name = data[name]
        for sheet in wb:
            if find == 1:
                break
            for row in sheet.iter_rows(min_col=1,max_col=4):
                if row[0].value == name and row[1].value == dif:
                    find = 1
                    base = ('%.1f'%row[3].value)
                    await ctx.send(f'{name} {dif}的定數是{base}')
                    if len(score)!=0:
                        await ctx.send(f'打{score[0]}分的R值是{rating(row[3].value,int(score[0]))}')

        if find == 0:
            await ctx.send('找不到這首歌')

    @commands.command()
    async def mai(self,ctx,name,dif,*other):
        type = None
        score = None

        def is_num(str):
            try:
                float(str)
                return True
            except ValueError:
                pass
            return False

        def rating(base,sc):
            if sc>=100.5:
                return round(base*22.4*sc/100-0.5)
            elif sc<100.5 and sc>=100:
                return round(base*21.6*sc/100-0.5)
            elif sc<100 and sc>=99.5:
                return round(base*21.1*sc/100-0.5)
            elif sc<99.5 and sc>=99:
                return round(base*20.8*sc/100-0.5)
            elif sc<99 and sc>=98:
                return round(base*20.3*sc/100-0.5)
            elif sc<98 and sc>=97:
                return round(base*20.0*sc/100-0.5)
            elif sc<97 and sc>=95:
                return round(base*16.8*sc/100-0.5)
            elif sc<95 and sc>=90:
                return round(base*15.2*sc/100-0.5)
            elif sc<90 and sc>=80:
                return round(base*13.6*sc/100-0.5)
            else:
                return 0

        if len(other)!=0:
            for item in other:
                if is_num(item):
                    score = float(item)
                else:
                    type = item


        wb = load_workbook("mai定數.xlsx")
        list = []
        with open('abbreviation.json','r',encoding='utf-8') as j:
            data = json.load(j)
            j.close
        if name in data:
            name = data[name]
        for sheet in wb:
            for row in sheet.iter_rows(min_col=1,max_col=5):
                if row[0].value == name and row[3].value == dif:
                    values = [cell.value for cell in row]
                    list.append(values)

        if len(list) == 2 and type == None:
            await ctx.send("這首歌有兩個譜面，你是要STD譜還是DX譜")

            def check(m):
                return (m.content == "DX" or m.content == "STD") and m.channel == ctx.channel
            msg = await self.bot.wait_for("message",check=check)
            type = msg.content
            for item in list:
                if item[2] == type:
                    base = ('%.1f'%item[4])
                    await ctx.send(f'{name} {dif}({type}譜)的定數是{base}')
                    if score!= None:
                        await ctx.send(f'{name} {dif}({type}譜)打{score}%的R值是{rating(item[4],score)}')
        elif len(list) == 2 and type != None:
            for item in list:
                if item[2] == type:
                    base = ('%.1f'%item[4])
                    await ctx.send(f'{name} {dif}({type}譜)的定數是{base}')
                    if score!= None:
                        await ctx.send(f'{name} {dif}({type}譜)打{score}%的R值是{rating(item[4],score)}')
        elif len(list) == 1:
            item = list[0]
            base = ('%.1f'%item[4])
            await ctx.send(f'{name} {dif}的定數是{base}')
            if score!= None:
                await ctx.send(f'{name} {dif}打{score}%的R值是{rating(item[4],score)}')
        else:
            await ctx.send('找不到這首歌')

async def setup(bot:commands.Bot) -> None:
    await bot.add_cog(Chumai(bot))