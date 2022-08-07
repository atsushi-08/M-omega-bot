import discord
from discord.ext import commands
from core.classes import cog_extension
from openpyxl import load_workbook

class Genshin(cog_extension):
    @commands.command()
    async def 原神(self,ctx,name,option):
        if(option == '突破'):
            list = []
            wb = load_workbook('evolve_material.xlsx')
            ws = wb.active
            for col in ws.iter_cols(max_col=1,max_row=52):
                i = 1
                find = 0
                for cell in col:
                    if(name == cell.value):
                        for row in ws.iter_rows(min_col=2,max_col=11,min_row=i,max_row=i):
                            for x in row:
                                tmp = x.value.replace("*","x")
                                list.append(tmp)
                        find = 1
                        break
                    if(find == 1):
                        break
                    i = i+1
            if(find == 0):
                await ctx.send("找不到這個人")
            else:
                embed = discord.Embed(
                    title = f'{name}的突破素材',
                    description = "從1等到90等的量",
                    color = ctx.author.color
                )
                list = '\n'.join(list)
                embed.add_field(value = list,name = "素材列表")
                await ctx.send(embed=embed)
        elif(option == '武器'):
            list = []
            wb = load_workbook('weapon.xlsx')
            ws = wb.active
            for col in ws.iter_cols(max_col=1,max_row=52):
                i = 1
                find = 0
                for cell in col:
                    if(name == cell.value):
                        for row in ws.iter_rows(min_col=2,min_row=i,max_row=i):
                            for x in row:
                                if(x.value == None):
                                    break
                                else:
                                    list.append(x.value)
                        find = 1
                        break
                    if(find == 1):
                        break
                    i = i+1
            if(find == 0):
                await ctx.send("找不到這個人")
            else:
                embed = discord.Embed(
                    title = f'{name}的推薦武器及理由',
                    color = ctx.author.color,
                    description = '參考一下就好啊臭ㄐㄐ'
                )
                for n in range(len(list)):
                    tmp = list[n].partition(':')
                    embed.add_field(name=tmp[0],value=tmp[2])
                await ctx.send(embed=embed)

                        

def setup(bot):
    bot.add_cog(Genshin(bot))