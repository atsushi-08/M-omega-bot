from openpyxl import load_workbook
import discord
from discord.ext import commands

class Chunithm(commands.Cog):
    def __init__(self,bot :commands.Bot) -> None:
        self.bot = bot
    
    @commands.command()
    async def 中二(self,ctx,name,dif):
        wb = load_workbook('中二定數.xlsx')
        find = 0
        for sheet in wb:
            if find == 1:
                break
            for row in sheet.iter_rows(min_col=1,max_col=4):
                if row[0].value == name and row[1].value == dif:
                    find = 1
                    await ctx.send(f'{name} {dif}的定數是{row[3].value}')
        if find == 0:
            await ctx.send('找不到這首歌')

async def setup(bot:commands.Bot) -> None:
    await bot.add_cog(Chunithm(bot))